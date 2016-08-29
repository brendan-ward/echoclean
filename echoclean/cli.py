import os
import sys
import logging
import time
import datetime
import re
import click
from collections import Counter
from csv import DictReader as CSV_DictReader

from openpyxl import Workbook

from echoclean.xlsx_dictreader import DictReader as XLSX_DictReader
from echoclean.ruleset import Ruleset

logger = logging.getLogger('echoclean')


def configure_logging(verbose):
    if verbose == 2:
        level = logging.DEBUG
    elif verbose == 1:
        level = logging.INFO
    else:
        level = logging.ERROR
    logging.basicConfig(stream=sys.stderr, level=level)


# For parsing sonobat .wav filename into a timestamp
FILENAME_RE = re.compile(
    '(?P<year>\d{4})(?P<month>\d{2})(?P<day>\d{2})_(?P<hour>\d{2})'
    '(?P<minute>\d{2})(?P<second>\d{2})_(?P<ms>\d{3})'
)


def _extract_sonobat_night(pass_filename):
    """
    Extract night from sonobat filename.
    Consider anything prior to noon as part of the previous night.
    """
    match = FILENAME_RE.search(pass_filename)
    if match:
        values = [int(v) for v in match.groups()]
        values[-1] = values[-1] * 1000  # convert ms to microseconds
        d = datetime.datetime(*values)
        if d.hour < 12:
            d = d - datetime.timedelta(days=1)
        return d.strftime('%m/%d/%Y')
    return None


@click.group()
def cli():
    pass


@cli.command(short_help='Apply the rules to the input data.')
@click.argument('rules', type=click.Path(exists=True))
@click.argument('data', type=click.Path(exists=True))
@click.argument('output', type=click.Path(dir_okay=False, writable=True, exists=False), required=False)
@click.option('-v', '--verbose', count=True, help='Verbose output')
def apply(rules, data, output, verbose):
    """Apply the rules to the input data."""

    configure_logging(verbose)

    try:
        data_ext = os.path.splitext(data)[1]
        if data_ext == '.xlsx':
            data_reader = XLSX_DictReader.from_file(data, prompt=True)
        elif data_ext in ('.txt', '.csv'):
            data_file = open(data)
            if data_ext == '.csv':
                delim = ','
            else:  # may be tab or comma delimited
                delim = '\t' if '\t' in data_file.read(1000) else ','
                data_file.seek(0)
            data_reader = CSV_DictReader(data_file, delimiter=delim)

        else:
            raise click.BadParameter('data file must be an XLSX, TXT, or CSV file',
                                     param='data', param_hint='data')

        rules_ext = os.path.splitext(rules)[1]
        if rules_ext == '.xlsx':
            rule_reader = XLSX_DictReader.from_file(rules, prompt=True)
        elif rules_ext in ('.txt', '.csv'):
            rules_file = open(rules)
            if rules_ext == '.csv':
                delim = ','
            else:  # may be tab or comma delimited
                delim = '\t' if '\t' in rules_file.read(1000) else ','
                rules_file.seek(0)
            rule_reader = CSV_DictReader(rules_file, delimiter=delim)

        else:
            raise click.BadParameter('rules file must be an XLSX, TXT, or CSV file',
                                     param='rules', param_hint='rules')

    except ValueError as e:
        print(e.message)
        raise click.Abort()

    if not output:
        output = '{0}_out.xlsx'.format(os.path.splitext(data)[0])

    elif not '.xlsx' in output:
        output += '.xlsx'

    # Extract out columns into criteria or new
    criteria_cols = []
    result_cols = []  # Any column in rule that doesn't exist in input is used for output when that rule is met
    for c in rule_reader.fieldnames:
        if c in data_reader.fieldnames:
            criteria_cols.append(c)
        else:
            result_cols.append(c)

    has_pass_filename = 'Filename' in data_reader.fieldnames

    print('Criteria columns: {}'.format(','.join(criteria_cols)))
    print('New columns added for results: {}'.format(','.join(result_cols)))

    # Parse rules
    start = time.time()
    rules = [rule for rule in rule_reader]
    ruleset = Ruleset(rules, result_cols)
    logger.debug('Parsed {} rules in {:.2f} seconds'.format(len(rules),
                                                            time.time() - start))
    logger.debug('Ruleset:\n{0}\n'.format(ruleset))

    # Setup output worksheet
    start = time.time()
    output_wb = Workbook(write_only=True)
    output_ws = output_wb.create_sheet(index=0, title='Classify Results')

    output_cols = list(result_cols)
    if has_pass_filename:
        output_cols += ['night']

    output_ws.append(output_cols + data_reader.fieldnames)

    empty_row = [''] * len(result_cols)

    inspect_idx = result_cols.index('Inspect')
    if inspect_idx >= 0:
        empty_row[inspect_idx] = 'Yes'

    print('\nClassifying passes')
    counter = 0
    counts = {k: dict() for k in result_cols}
    nights = {}
    classified = Counter()

    for i, row in enumerate(data_reader):
        logger.info('classifying row #{0}'.format(i))

        # Classify row
        result = ruleset.test(row)
        if result:
            output_row = result
            for index, key in enumerate(result_cols):
                value = result[index]
                if value not in counts[key]:
                    counts[key][value] = 0
                counts[key][value] += 1
        else:
            output_row = list(empty_row)

        classified.update([bool(result)])

        if has_pass_filename:
            filename = os.path.split(row['Filename'])[1]
            night = _extract_sonobat_night(filename) or ''
            if night:
                if night not in nights:
                    nights[night] = Counter()
                nights[night].update([bool(result)])

            output_row.append(night)

        # row.values() does not preserve order when using csv
        orig_values = [row[k] for k in data_reader.fieldnames]
        output_row.extend(orig_values)
        output_ws.append(output_row)

        counter += 1
        # if counter > 2:  # FIXME
        #     break



    # Report overall hits and misses
    output_ws = output_wb.create_sheet(index=1, title='Classification Summary')
    output_ws.append(['Classified', 'Count'])
    for result in (True, False):
        output_ws.append(['Yes' if result else 'No', classified[result]])
    output_ws.append(['Total', sum(classified.values())])

    if has_pass_filename and len(nights):
        # Report hits and misses for each night
        output_ws = output_wb.create_sheet(index=2, title='Night Classification Summary')
        output_ws.append(
            ['Night', 'Rows Classified', 'Rows Not Classified', 'Total Rows'])
        nights_keys = list(nights.keys())
        nights_keys.sort()
        for night in nights_keys:
            output_ws.append(
                [night] + [nights[night][value] for value in (True, False)] + [
                    sum(nights[night].values())])

    # Create a summary sheet for all values in classified output
    for i, key in enumerate(result_cols):
        if len(counts[key].keys()) > 1:
            output_ws = output_wb.create_sheet(index=i + 3, title='{0} Summary'.format(key))
            output_ws.append(['Value', 'Rows Classified'])
            values = [x for x in counts[key].keys() if x]
            if len(set(type(x) for x in values)) == 1:
                # cannot sort if values is mixed types
                values.sort()
            if None in counts[key].keys():
                values = [None] + values
            for value in values:
                output_ws.append([value, counts[key][value]])

    output_wb.save(filename=output)


    print('\nEvaluated {} passes in {:.2f} seconds'.format(counter, time.time() - start))
    for result in (True, False):
        print('{1} {0}'.format('matched a rule' if result else 'did not match any rules', classified[result]))