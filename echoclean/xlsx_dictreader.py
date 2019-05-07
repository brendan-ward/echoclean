import os
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from six.moves import input


class DictReader(object):
    def __init__(self, worksheet):
        self.fieldnames = []
        column_idx = []
        row_iter = worksheet.iter_rows()

        for cell in next(row_iter):
            if not cell.value:  # Any blank column is beyond range of data
                break
            self.fieldnames.append(cell.value)
            column_idx.append(cell.column)

        self.column_range = (column_idx[0], column_idx[-1])

        # Setup rows iter
        min_col, min_row, max_col, max_row = range_boundaries('{0}2:{1}{2}'.format(
            self.column_range[0],
            self.column_range[1],
            worksheet.max_row
        ))
        self._iter_rows = worksheet.iter_rows(
            min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
        )

    def __iter__(self):
        return self

    def __next__(self):
        row = next(self._iter_rows)
        return OrderedDict([(k, v) for k, v in zip(self.fieldnames, [r.value for r in row])])

    @classmethod
    def from_file(cls, filename, index=None, prompt=False):
        """Read sheet from filename
        If index is not specified, the active sheet will be used.
        If multiple sheets are present and prompt is true, the user will be
        prompted to enter the index.
        """

        workbook = load_workbook(filename, data_only=True)
        worksheet = workbook.active
        if index is not None:
            worksheet = workbook.worksheets[index]

        elif prompt and len(workbook.worksheets) > 1:
            print('\n-------------------------------------------------------------\nMultiple sheets found:')
            valid_idx = []
            for i, ws in enumerate(workbook.worksheets):
                if ws.max_row:
                    print('{0})  {1}{2}'.format(i, ws.title,
                        ' (default)' if i == workbook._active_sheet_index else ''))
                    valid_idx.append(i)
            print('Choose sheet number and press ENTER or leave blank for default:')
            index = input('\n>  ')
            print('')

            if index != '':
                index = int(index)
                if not index in valid_idx:
                    raise ValueError('Invalid Sheet Number')

                worksheet = workbook.worksheets[index]

        if len(workbook.worksheets) > 1:
            print('{0}: using worksheet {1}'.format(os.path.split(filename)[1], worksheet.title))
        return cls(worksheet)